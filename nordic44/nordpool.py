"""Module for getting data from nordpool."""
from ftplib import FTP
from openpyxl import Workbook
from pandas import ExcelFile


def fix_digit(x):
    """Function to write week or year with 2 digits.
    Args:
        x: week or year as int
    """
    x_str = str(x)
    if len(x_str) == 1:
        return "0" + x_str
    else:
        return x


def read_excel(fname, skiprows=0, header=None):
    """Read excel into dict.
    Args:
        fname: name of excel file
        skip_rows: If any rows should be skipped
    """
    xls = ExcelFile(fname)
    if header:
        parse_cols = [1]
    else:
        parse_cols = None

    df = xls.parse(xls.sheet_names[0], skiprows=1,
                   parse_cols=parse_cols)

    # Fix keys
    temp = df.to_dict()
    for key in temp:
        new_key = key.replace(" - ", "_")
        temp[new_key] = temp.pop(key)
    # Stupid hack for Finland
    if header:
        temp[header] = temp.pop(temp.keys()[0])

    return temp


class NordPool():
    """Class for getting the data from NordPool.

    Args:
        user: User to login at NordPool
        passwd: password for NordPool
    """

    def __init__(self, date):
        """Constructor."""
        # Codes used in NordPool
        self.code = ["PS", "FB", "UT"]
        # Have to use list comprehension otherwise they point to
        # the same lists
        self.Norway = {key: dict() for key in self.code}
        self.Sweden = {key: dict() for key in self.code}
        self.Finland = {key: dict() for key in self.code}

        # Maybe I will remove the Norway, Sweden and Finland dicts later
        self.data = dict()
        self.data["NO"] = self.Norway
        self.data["SE"] = self.Sweden
        self.data["FI"] = self.Finland

        self.countries = ["Norway", "Sweden", "Finland"]

        self.date = date

    def add_content(self, country, key, code, data):
        """Method that reads in data and update the keys in the dicts.
        Args:
            country: The country to do the operation for
            key: Area code for instance NO1
            code: NordPool code for instance PS or FB
            data: data to add to the dict
        """
        x = getattr(self, country)
        x[code][key] = map(int, filter(lambda val: val != "", data))

    def read_data_from_ftp(self, user, passwd):
        """Function to read in the data from NordPool"""
        ftp = FTP("ftp.nordpoolspot.com")

        ftp.login(user=user, passwd=passwd)
        ftp.cwd("Operating_data")

        fnames = ["pono", "pose", "pofi"]

        week = self.date.isocalendar()[1]

        # The first days of a year can be in the file of the 53 week
        # of the previous year
        if week == 53:
            year_str = str(self.date.year-1)
        else:
            year_str = str(self.date.year)

        
        week_day = str(self.date.weekday()+1)

        week_str = fix_digit(str(week))
        year_2_str = fix_digit(year_str[2:])

        back = ".."
        i = 0
        for country in self.countries:
            content = []
            ftp.cwd(country)

            if self.date.year < 2016:
                ftp.cwd(year_str)
                back = "../.."

            ftp.retrlines(
                "RETR " + fnames[i] + year_2_str + week_str + ".sdv",
                content.append)

            temp = []
            for line in content:
                temp[:] = line.split(";")
                if len(temp) > 3 and temp[4] == week_day:
                    if temp[0] == self.code[0] and temp[1] == self.code[0][0]:
                        self.add_content(country, temp[6], self.code[0],
                                         temp[7:])
                    if temp[0] == self.code[1] and temp[1] == self.code[1][0]:
                        self.add_content(country, temp[6], self.code[1],
                                         temp[7:])
                    if temp[0] == self.code[2] and temp[1] == self.code[2][0]:
                        self.add_content(country, temp[6], self.code[2],
                                         temp[7:])
            ftp.cwd(back)
            i = i + 1

    def read_data_from_excel(self, no_prod, se_prod, fi_prod,
                             no_con, se_con, fi_con,
                             no_ex, se_ex, fi_ex):
        """ Method for reading nordpool data from excel
        Args:
            no_prod: Excel file containing Norway's production data
            se_prod: Excel file containing Sweden's production data
            fi_prod: Excel file containing Finland's production data
            no_con: Excel file containing Norway's consumption data
            se_con: Excel file containing Sweden's consumption data
            fi_con: Excel file containing Finland's consumption data
            no_ex: Excel file containing Norway's exchange data
            se_ex: Excel file containing Sweden's exchange data
            fi_ex: Excel file containing Finland's exchange data
        """
        self.data["NO"]["PS"] = read_excel(no_prod, skiprows=1)
        self.data["SE"]["PS"] = read_excel(se_prod, skiprows=1)
        self.data["FI"]["PS"] = read_excel(fi_prod, skiprows=1,
                                           header="FI")

        self.data["NO"]["FB"] = read_excel(no_con, skiprows=1)
        self.data["SE"]["FB"] = read_excel(se_con, skiprows=1)
        self.data["FI"]["FB"] = read_excel(fi_con, skiprows=1,
                                           header="FI")

        self.data["NO"]["UT"] = read_excel(no_ex, skiprows=1)
        self.data["SE"]["UT"] = read_excel(se_ex, skiprows=1)
        self.data["FI"]["UT"] = read_excel(fi_ex, skiprows=1)

    def write_data_to_excel(self):
        """Write the data to excel in the iTesla format."""
        fnames = ["Procution", "Consumption", "Exchange"]
        country_codes = ["NO", "SE", "FI"]
        for cc, country in zip(country_codes, self.countries):
            temp = getattr(self, country)
            for code, name in zip(self.code, fnames):
                x = temp[code]
                wb = Workbook()
                ws = wb.active
                ws.append([name + " in MWh/h"])
                ws.append([self.date])

                # Add area headers
                column = 2
                for key in x:
                    ws.cell(row=2, column=column).value = key
                    column = column + 1
                for i in range(3, 28):
                    if i == 27:
                        value = "SUM"
                    else:
                        value = str(i-3) + " - " + str(i-2)
                    ws.cell(row=i, column=1).value = value

                column = 2
                row = 3
                for key in x:
                    for data in x[key]:
                        value = data
                        ws.cell(row=row, column=column).value = value
                        row = row + 1
                    column = column + 1
                    row = 3
                wb.save(name + "_" + cc + ".xlsx")

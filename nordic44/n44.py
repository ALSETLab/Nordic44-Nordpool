"""Functions for updating the PSS/E data sets."""
import os
import sys
import math
import warnings
import psspy
import redirect
import openpyxl
from openpyxl.styles import Alignment
from collections import namedtuple
from collections import OrderedDict

# Define enums containing the mapping
ExchangeLoad = namedtuple("ExchangeLoad", "bus, i, area, areas, pf")
AreaInfo = namedtuple("AreaInfo", "number, bus, pf, pos, neg")


class N44(object):
    """Class mapping nordpool and N44
    Args:
        basecase: name of the PSS/E base case
        data: Nordpool data dictionary
    """
    def __init__(self, data, basecase=None):
        self.data = data
        if not basecase:
            cwd = os.path.dirname(os.path.abspath(__file__))
            encoding = sys.getfilesystemencoding()
            self.basecase = os.path.join(
                cwd.encode(encoding),
                "models",
                "N44_BC.sav")
        else:
            self.basecase = basecase

        self.wb = None
        self.sheet = None
        self.to_excel = True

        # In the future these mappings could be input
        # Mapping of exchanges represented as loads
        pf = 0.95

        self.ex_as_load = ([
            ExchangeLoad(3020, '1', "SE3", "SE3_FI", pf),
            ExchangeLoad(3360, '1', "SE3", "SE3_DK1", pf),
            ExchangeLoad(5610, '1', "NO2", "NO_DK", 0.99),
            ExchangeLoad(5620, '1', "NO2", "NO_NL", pf),
            ExchangeLoad(6701, '1', "NO4", "NO_FI", pf),
            ExchangeLoad(6701, '3', "NO4", "NO_RU", pf),
            ExchangeLoad(7000, '6', "FI1", "FI_SE3", pf),
            ExchangeLoad(7010, '1', "FI1", "FI_RU", pf),
            ExchangeLoad(7020, '1', "FI1", "FI_EE", pf),
            ExchangeLoad(7100, '3', "FI1", "FI_NO", pf),
            ExchangeLoad(8500, '4', "SE4", "SE4_DK2", pf),
            ExchangeLoad(8600, '1', "SE4", "SE_DE", pf),
            ExchangeLoad(8700, '1', "SE4", "SE_PL", pf)])

        # Mapping of interarea exchanges
        self.area_info = OrderedDict(
            [("NO1",
              AreaInfo(11, 5500, pf,
                       pos=[],
                       neg=["NO1_SE3", "NO1_NO2", "NO1_NO5", "NO1_NO3"])),
             ("NO2",
              AreaInfo(12, 5600, pf,
                       pos=["NO1_NO2"],
                       neg=["NO2_NO5"])),
             ("NO3",
              AreaInfo(13, 6500, pf,
                       pos=["NO1_NO3"],
                       neg=["NO3_SE2", "NO3_NO4"])),
             ("NO4",
              AreaInfo(14, 6700, pf,
                       pos=["NO3_NO4"],
                       neg=["NO4_SE1", "NO4_SE2"])),
             ("NO5",
              AreaInfo(15, 5300, pf,
                       pos=["NO1_NO5", "NO2_NO5"],
                       neg=[])),
             ("SE1",
              AreaInfo(21, 3115, pf,
                       pos=[],
                       neg=["SE1_FI", "SE1_NO4", "SE1_SE2"])),
             ("SE2",
              AreaInfo(22, 3249, pf,
                       pos=["SE1_SE2"],
                       neg=["SE2_NO3", "SE2_NO4", "SE2_SE3"])),
             ("SE3",
              AreaInfo(23, 3500, pf,
                       pos=["SE2_SE3"],
                       neg=["SE3_NO1", "SE3_SE4"])),
             ("SE4",
              AreaInfo(24, 8500, pf,
                       pos=["SE3_SE4"],
                       neg=[])),
             ("FI",
              AreaInfo(31, 7000, pf,
                       pos=[],
                       neg=["FI_SE1"]))])

    def update_raw_files(self, to_excel=True, out_dir=None):
        """Function for updating the psse case file.
        Args:
            to_excel(default=True): If a summary should be written to excel
            out_dir: The directory where the results are stored
        """
        if not out_dir:
            out_dir = os.getcwd()

        redirect.psse2py()
        psspy.throwPsseExceptions = True
        nbuses = 50000  # max no of buses
        ierr = psspy.psseinit(nbuses)
        psspy.case(self.basecase)

        if to_excel:
            self.create_excel_sheet()
            self.to_excel = True
        else:
            self.sheet = None

        for i, col in zip(range(0, 24),
                          range(2, 2+24*3, 3)):
            # Represent HVDC links as load and some other exchanges as well
            print('Changing additional loads...')

            row = 15
            for load in self.ex_as_load:
                self.load_change(load, i,
                                 to_excel, row, col)
                row = row + 1

            print('Changing interarea exchanges...')
            row = 3
            for area, info in self.area_info.items():
                country = area[0:2]
                # Changing interarea exchanges
                exchange = self.calculate_exchange(info, area, i)
                self.area_data(
                    info.number,
                    info.bus,
                    exchange,
                    area,
                    row, col+2)
                # Changing areas production and consumption
                self.change_prod_con(
                    info.number,
                    self.data[country]["PS"][area][i],
                    self.data[country]["FB"][area][i],
                    info.pf,
                    tol=4, row=row, column=col)
                row = row + 1

            print('Changes completed...')
            # Save the raw file to convert to CIM
            psspy.rawd_2(0, 1, [0, 0, 1, 0, 0, 0, 0], 0, "Snap_before_PF.raw")
            b = 'h' + str((col-1)/3) + '_before_PF.raw'
            os.rename("Snap_before_PF.raw", b)
            # Solve the power flow
            psspy.fnsl([1, 2, 0, 0, 1, 0, 0, 0])
            ival = psspy.solved()  # flag to check power flow convergence
            if ival == 0:
                print('Convergence')
                if self.to_excel:
                    self.sheet.cell(row=42, column=col).value = 'Convergence'
                
                temp_fname = os.path.join(out_dir, "temp.sav")
                print(temp_fname)

                psspy.save(temp_fname)  # save temporarily the solved case
                psspy.case(temp_fname)  # set the saved case as current case

                # save the raw file to convert to CIM
                raw_fname = os.path.join(out_dir, "Snap_after_PF.raw")
                psspy.rawd_2(0, 1, [0, 0, 1, 0, 0, 0, 0], 0,
                             raw_fname)
                b = os.path.join(out_dir, 'h' + str(i) + '_after_PF.raw')
                os.rename(raw_fname, b)

                if self.to_excel:
                    # Merge cells
                    self.sheet.merge_cells(start_row=1, start_column=col,
                                           end_row=1, end_column=col+2)

                    self.sheet.cell(row=2, column=col).alignment = (
                        Alignment(wrapText=True))
                    self.sheet.cell(row=2, column=col+1).alignment = (
                        Alignment(wrapText=True))
                    self.sheet.cell(row=2, column=col+2).alignment = (
                        Alignment(wrapText=True))
                    self.sheet.cell(row=14, column=col).alignment = (
                        Alignment(wrapText=True))
                    self.sheet.cell(row=14, column=col+1).alignment = (
                        Alignment(wrapText=True))
                    self.sheet.cell(row=30, column=col).alignment = (
                        Alignment(wrapText=True))
                    self.sheet.cell(row=30, column=col+1).alignment = (
                        Alignment(wrapText=True))

                    # Headers for data from nordpool
                    self.sheet.cell(row=1, column=col).value = (
                        'hour ' + str(i))
                    self.sheet.cell(row=2, column=col).value = (
                        'Scheduled\nProduction\n[MWh]')
                    self.sheet.cell(row=2, column=col+1).value = (
                        'Scheduled\nConsumption\n[MWh]')
                    self.sheet.cell(row=2, column=col+2).value = (
                        'Scheduled\nExchange\n[MWh]')

                    # Headers for exchanges represented as loads
                    self.sheet.cell(row=14, column=col).value = (
                        'Active Power\n[MW]')
                    self.sheet.cell(row=14, column=col+1).value = (
                        'Reactive Power\n[MW]')

                    # Headers for results after PSS/E
                    self.sheet.cell(row=30, column=col).value = (
                        'PSSE\nProduction\n[MWh]')
                    self.sheet.cell(row=30, column=col+1).value = (
                        'PSSE\nConsumption\n[MWh]')
                    self.sheet.cell(row=30, column=col+2).value = (
                        'PSSE\nExchange\n[MWh]')

                    row = 31
                    for _, info in self.area_info.items():
                        # to get the area production complex power
                        ierr = psspy.ardat(info.number, 'GEN')
                        self.sheet.cell(row=row, column=col).value = (
                            round(ierr[1].real, 0))

                        # to get the area consumption complex power
                        ierr = psspy.ardat(info.number, 'LOAD')
                        self.sheet.cell(row=row, column=col+1).value = (
                            round(ierr[1].real, 0))
                        row += 1

                    # to get the value of the areas active power interchange
                    ierr, intch = psspy.aareareal(-1, 1, 'PINT')
                    for r in range(0, len(intch[0])):
                        self.sheet.cell(
                            row=31+r,
                            column=col+2).value = round(intch[0][r].real, 0)

                    # limits check
                    ierr, busvoltages = psspy.abusreal(sid=-1, string="PU")
                    if any(x < 0.95 or x > 1.05 for x in busvoltages[0]):
                        self.sheet.cell(row=43, column=col).value = (
                            'Bus voltage problem')
                    ierr, machPGen = psspy.amachreal(sid=-1, string="PGEN")
                    ierr, machPMax = psspy.amachreal(sid=-1, string="PMAX")
                    ierr, machPMin = psspy.amachreal(sid=-1, string="PMIN")
                    ierr, machQGen = psspy.amachreal(sid=-1, string="QGEN")
                    ierr, machQMax = psspy.amachreal(sid=-1, string="QMAX")
                    ierr, machQMin = psspy.amachreal(sid=-1, string="QMIN")
                    ierr, machS = psspy.amachreal(sid=-1, string="MVA")
                    ierr, machMbase = psspy.amachreal(sid=-1, string="MBASE")
                    for l in range(0, len(machPGen[0])):
                        if (machPGen[0][l] <= machPMin[0][l] or
                                machPGen[0][l] >= machPMax[0][l]):
                            self.sheet.cell(
                                row=45,
                                column=col).value = (
                                    'Generator active power output problem')
                for m in range(0, len(machQGen[0])):
                    if (machQGen[0][m] <= machQMin[0][m] or
                            machQGen[0][m] >= machQMax[0][m]):
                        self.sheet.cell(row=46, column=col).value = (
                            'Generator reactive power output problem')
                        break
                for n in range(0, len(machS[0])):
                    if machS[0][n] >= machMbase[0][n]:
                        self.sheet.cell(row=47, column=col).value = (
                            'Generator overloading problem')
                        break
                ierr, brflowA = psspy.aflowreal(sid=-1, string="PCTCORPRATEA")
                if any(x >= 100 for x in brflowA[0]):
                    self.sheet.cell(row=48, column=col).value = (
                        'Branch overloading problem (Rate A)')
                ierr, brflowB = psspy.aflowreal(sid=-1, string="PCTCORPRATEB")
                if any(x >= 100 for x in brflowB[0]):
                    self.sheet.cell(row=48, column=col).value = (
                        'Branch overloading problem (Rate B)')
                ierr, brflowC = psspy.aflowreal(sid=-1, string="PCTCORPRATEC")
                if any(x >= 100 for x in brflowC[0]):
                    self.sheet.cell(row=48, column=col).value = (
                        'Branch overloading problem (Rate C)')
            else:
                print('No convergence')
                self.sheet.cell(row=43, column=col).value = 'No convergence'

        psspy.close_powerflow()

        # save the Excel file with all data
        self.wb.save(os.path.join(out_dir, 'PSSE_in_out.xlsx'))
        os.remove(temp_fname)

    def change_prod_con(self, areas, prod, con, pf, tol=4,
                        row=None, column=None):
        """Wrapper function to change load and production
        Args:
            areas: area number
            prod: production
            con: consumption
            pf: power factor
            tol: tolerance for round
            row: which row to write to
            column: which column to write to
        """
        psspy.bsys(sid=0, numarea=1, areas=[areas])
        psspy.scal_2(0, 0, 0,
                     [0, 1, 1, 1, 0],
                     [con, prod,
                      0.0, 0.0, 0.0, 0.0,
                      round(con*math.tan(math.acos(pf)), tol)])
        if self.to_excel:
            self.sheet.cell(row=row, column=column).value = prod
            self.sheet.cell(row=row, column=column+1).value = con

    def load_change(self, load, hour, tol=4,
                    row=None, column=None):
        """Wrapper function for changing loads in PSS/E
        Args:
            load: LoadExchange
            hour: Time of day
            tol: Tolerance of round
            row: which row to write to
            column:which column to write to
        """
        country = load.area[0:2]
        try:
            realar = -self.data[country]["UT"][load.areas][hour]
        except IndexError:
            realar = 0
            warnings.warn("Missing data filling in zero")
            with open("warnings.txt", "a") as warn:
                warn.write("Missing data for " +
                           load.areas + " at hour " +
                           str(hour))
        except KeyError:
            # Sweden hack
            temp = load.area + load.areas[2:]
            realar = -self.data[country]["UT"][temp][hour]

        realar2 = round(realar*math.tan(math.acos(load.pf)), tol)

        # Update the load
        if country == "FI":
            area = "FI"
        else:
            area = load.area
        self.data[country]["FB"][area][hour] += (
            realar)

        if self.to_excel:
            self.sheet.cell(row=row, column=column).value = realar
            self.sheet.cell(row=row, column=column+1).value = realar2

        psspy.load_chng_4(load.bus, load.i, realar1=realar, realar2=realar2)

    def area_data(self, i, intgar, realar, arname,
                  row=None, column=None):
        """Wrapper function for changing loads in PSS/E
        Args:
            i: Area number
            intgar: In this case the swing bus
            realar: See PSS/E documentation
            arname: Name of the area
            row: which row to write to
            column:which column to write to
        """
        if self.to_excel:
            self.sheet.cell(row=row, column=column).value = realar
        psspy.area_data(i, intgar, realar1=realar, arname=arname)

    def calculate_exchange(self, info, area, hour):
        """Calculates area exchanges
        Args:
            info: Areainfo
            area: name of area
            hour: Hour index
        Outputs:
            exchange
        """
        exchange = 0
        temp = self.data[area[0:2]]["UT"]
        for pos in info.pos:
            try:
                exchange += temp[pos][hour]
            except KeyError:
                exchange -= temp[flip_area(pos)][hour]
        for neg in info.neg:
            try:
                exchange -= temp[neg][hour]
            except KeyError:
                exchange += temp[flip_area(neg)][hour]
        return exchange

    def create_excel_sheet(self):
        """Function to set up the summary excel sheet.
        """

        self.wb = openpyxl.Workbook()
        sheet = self.wb.active
        # Row and column headings
        sheet['A3'] = 'NO1'
        sheet['A4'] = 'NO2'
        sheet['A5'] = 'NO3'
        sheet['A6'] = 'NO4'
        sheet['A7'] = 'NO5'
        sheet['A8'] = 'SE1'
        sheet['A9'] = 'SE2'
        sheet['A10'] = 'SE3'
        sheet['A11'] = 'SE4'
        sheet['A12'] = 'FI1'
        sheet['A14'] = 'Additional loads'
        sheet['A15'] = 'Bus 3020, area SE3, HVDC link SE3-FI'
        sheet['A16'] = 'Bus 3360, area SE3, HVDC link SE3-DK1'
        sheet['A17'] = 'Bus 5610, area NO2, HVDC link NO-DK'
        sheet['A18'] = 'Bus 5620, area NO2, HVDC link NO-NL'
        sheet['A19'] = 'Bus 6701, area NO4, exchange NO-FI'
        sheet['A20'] = 'Bus 6701, area NO4, exchange NO-RU'
        sheet['A21'] = 'Bus 7000, area FI1, HVDC link FI-SE3'
        sheet['A22'] = 'Bus 7010, area FI1, HVDC link FI-RU'
        sheet['A23'] = 'Bus 7020, area FI1, HVDC link FI-EE'
        sheet['A24'] = 'Bus 7100, area FI1, exchange FI-NO'
        sheet['A25'] = 'Bus 8500, area SE4, link SE4-DK2'
        sheet['A26'] = 'Bus 8600, area SE4, HVDC link SE-DE'
        sheet['A27'] = 'Bus 8700, area SE4, HVDC link SE-PL'
        sheet['A30'] = 'Results after Power Flow in PSSE'
        sheet['A31'] = 'NO1'
        sheet['A32'] = 'NO2'
        sheet['A33'] = 'NO3'
        sheet['A34'] = 'NO4'
        sheet['A35'] = 'NO5'
        sheet['A36'] = 'SE1'
        sheet['A37'] = 'SE2'
        sheet['A38'] = 'SE3'
        sheet['A39'] = 'SE4'
        sheet['A40'] = 'FI1'
        self.sheet = sheet


def flip_area(code):
    """Flip area string"""
    return code[4:] + "_" + code[0:3]
